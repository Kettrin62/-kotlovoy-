from os import sep

from catalog.permissions import IsAdminOrReadOnly
from django.http.response import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.serializers import ValidationError

from kotlovoy62.settings import (CANT_DELETE_STATUS, CUSTOM_SETTINGS_DRF,
                                 MEDIA_ROOT)

from .models import Delivery, Order, OrderHasElement, OrderStatus, Payment
from .permissions import UserGetAndCreateOnlyOrAdmin
from .serializers import (DeliverySerializer, OrderSerializer,
                          OrderStatusSerializer, PaymentSerializer)


class OrderSetPagination(PageNumberPagination):
    page_size_query_param = 'limit'
    page_size = CUSTOM_SETTINGS_DRF.get('PAGE_SIZE_ORDERS')


class OrderViewSet(viewsets.ModelViewSet):
    http_method_names = ('get', 'post', 'patch', 'delete',)
    queryset = Order.objects.all()
    search_fields = ('number', 'email', 'phoneNumber', 'order_sum')
    filterset_fields = ('status',)
    serializer_class = OrderSerializer
    permission_classes = (UserGetAndCreateOnlyOrAdmin,)
    pagination_class = OrderSetPagination

    def get_queryset(self):
        if self.request.user.is_staff:
            return Order.objects.all()
        return self.request.user.orders.all()

    def perform_create(self, serializer):
        try:
            serializer.save(
                elements={'elements': (self.request.data['elements'])},
                delivery={'delivery': (self.request.data['delivery'])},
                # payment={'payment': (self.request.data['payment'])},
            )
        except KeyError as err:
            raise ValidationError(
                {
                    'message': [
                        'В запросе не передан параметр: {}'.format(err)
                    ]
                }
            )

    def perform_update(self, serializer):
        order = self.get_object()
        try:
            serializer.save(
                instance=order,
                elements={'elements': (self.request.data['elements'])},
                status={'status': (self.request.data['status'])},
                delivery={'delivery': (self.request.data['delivery'])},
                # payment={'payment': (self.request.data['payment'])},
            )
        except KeyError as err:
            raise ValidationError(
                {
                    'message': [
                        'В запросе не передан параметр: {}'.format(err)
                    ]
                }
            )

    @action(
        ['get'], detail=True,
        permission_classes=(IsAuthenticated,)
    )
    def cancel_order(self, request, pk):
        order = get_object_or_404(Order, pk=pk)

        if request.user != order.user and not request.user.is_superuser:
            return Response(
                {
                    'message': [
                        'У вас недостаточно прав для выполнения данного '
                        'действия.'
                    ]
                },
                status=status.HTTP_403_FORBIDDEN
            )

        order_elements = list(OrderHasElement.objects.filter(order=order))
        for item in order_elements:
            cnt = item.amount
            item.element.stock += cnt
            item.element.save()
        order_cancelled = OrderStatus.objects.get(status='отменённый заказ')
        order.status = order_cancelled
        order.save()

        return Response(
            {
                'order': [
                    f'Заказ {order.number} отменён'
                ]
            },
            status=status.HTTP_200_OK
        )

    @action(
        ['get'], detail=True,
        permission_classes=(IsAuthenticated,)
    )
    def get_report(self, request, pk):
        order = get_object_or_404(Order, pk=pk)

        if request.user != order.user and not request.user.is_superuser:
            return Response(
                {
                    'message': [
                        'У вас недостаточно прав для выполнения данного '
                        'действия.'
                    ]
                },
                status=status.HTTP_403_FORBIDDEN
            )

        order_elements = list(OrderHasElement.objects.filter(order=order))
        order_data = {}
        cnt = 1
        discount = order.discount
        created = order.created.strftime('%d.%m.%Yг.')
        for item in order_elements:
            amount = item.amount
            article = item.element.article
            title = item.element.title
            start_price = item.price
            cur_price = item.cur_price
            order_data[cnt] = {
                "amount": amount,
                "article": article,
                "title": title,
                "start_price": start_price,
                "cur_price": cur_price,
            }
            cnt += 1

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=report.xls'

        wb = load_workbook(
            MEDIA_ROOT + sep + 'pattern.xlsx', read_only=False
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        sheet = wb.active
        sheet.cell(row=1, column=6).value = (
            f'Ведомость заказа № {order.number}'
        )
        sheet.cell(row=2, column=7).value = (
            f'от {created}'
        )
        row = 5
        for num, i_data in order_data.items():
            column = 2
            sheet.cell(row=row, column=column).value = num
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = i_data.get('article')
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = i_data.get('title')
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = i_data.get('amount')
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = (
                i_data.get('start_price')
            )
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = discount
            sheet.cell(row=row, column=column).border = thin_border
            column += 1
            sheet.cell(row=row, column=column).value = (
                i_data.get('cur_price') * i_data.get('amount')
            )
            sheet.cell(row=row, column=column).border = thin_border
            row += 1
        wb.save(response)
        return response


class DeliveryViewSet(viewsets.ModelViewSet):
    http_method_names = ('get', 'post', 'patch', 'delete',)
    queryset = Delivery.objects.all()
    serializer_class = DeliverySerializer
    permission_classes = (IsAdminOrReadOnly,)


class PaymentViewSet(viewsets.ModelViewSet):
    http_method_names = ('get', 'post', 'patch', 'delete',)
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = (IsAdminOrReadOnly,)


class OrderStatusViewSet(viewsets.ModelViewSet):
    http_method_names = ('get', 'post', 'patch', 'delete',)
    queryset = OrderStatus.objects.all()
    serializer_class = OrderStatusSerializer
    permission_classes = (IsAdminOrReadOnly,)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.status in CANT_DELETE_STATUS:
            return Response(
                {
                    'message': [
                        'Удаление предустановленных статусов запрещено!'
                    ]
                },
                status=status.HTTP_403_FORBIDDEN
            )
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_update(self, serializer):
        status = self.get_object()
        if status.status in CANT_DELETE_STATUS:
            raise ValidationError(
                {
                    'message': [
                        'Изменение предустановленных статусов запрещено!'
                    ]
                }
            )
        serializer.save(instance=status)
